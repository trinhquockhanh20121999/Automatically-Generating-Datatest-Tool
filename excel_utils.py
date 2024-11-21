import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
start_col = 0
start_row = 0

def find_table_start(file_path, header_fill_color="FFFFC000", sheetname=None):
    # Mở file Excel
    workbook = openpyxl.load_workbook(file_path)
 
    # Chọn sheet theo tên (sheetname), nếu không có sheetname thì sử dụng sheet mặc định
    if sheetname:
        sheet = workbook[sheetname]
    else:
        sheet = workbook.active  # Dùng sheet mặc định nếu không có sheetname

    # Duyệt qua các ô trong sheet để tìm ô có màu nền tương ứng
    for row in sheet.iter_rows():
        for cell in row:
            # Kiểm tra nếu ô có màu nền trùng với màu header
            if not cell.fill or not cell.fill.fgColor:
                continue
            
            color = cell.fill.fgColor.rgb  # Mã màu Hex của ô
            if color == header_fill_color:
                # Trả về vị trí ô đầu tiên của header
                return cell.coordinate
    return None  # Nếu không tìm thấy ô có màu nền tương ứng

def read_table_data(file_path, header_fill_color="FFFFC000", sheetname=None):
    global start_col
    global start_row

    # Gọi hàm find_table_start để xác định ô bắt đầu của bảng
    start_cell = find_table_start(file_path, header_fill_color, sheetname)
    
    if start_cell is None:
        raise ValueError("Không tìm thấy vị trí bắt đầu của bảng dữ liệu.")

    # Chuyển đổi tọa độ ô (ví dụ: 'A1') thành chỉ số hàng và cột
    start_col = openpyxl.utils.column_index_from_string(start_cell[0]) - 1
    start_row = int(start_cell[1:]) - 1

    # Đọc dữ liệu từ file Excel bắt đầu từ vị trí của bảng
    df = pd.read_excel(file_path, sheet_name=sheetname, header=start_row, engine='openpyxl')
    df = df.iloc[:, start_col:]
    df = df.fillna("")

    return df

def write_data_to_excel(file_path, col, row, data, sheetname=None):
    global start_row
    global start_col
    # Mở tệp Excel
    wb = openpyxl.load_workbook(file_path)
    
    # Chọn sheet nếu có tên sheet được cung cấp, nếu không thì chọn sheet đầu tiên
    if sheetname:
        sheet = wb[sheetname]
    else:
        sheet = wb.active
    print(f'{row} {col} {start_row} {start_col}')
    data_string = str(data)
    # Ghi giá trị vào ô
    sheet.cell(row=start_row + col + 2, column=start_col + row + 1, value=data_string)
    
    # Lưu tệp Excel
    wb.save(file_path)